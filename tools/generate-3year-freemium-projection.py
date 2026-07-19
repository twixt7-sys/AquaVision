"""
AquaVision 3-Year Freemium Projection
Pricing: Free | Premium P299/mo | Enterprise P1,499/mo
Target: Year 2 breakeven | Grounded in TAM/SAM/SOM research
"""

from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

OUT = Path(__file__).resolve().parent.parent / "exports"
OUT.mkdir(exist_ok=True)

# --- Market anchors (from research) ---
MARKET = {
    "tam_software_usd_b": 1.5,  # midpoint aqua software / digital (~USD 1-2B)
    "tam_precision_usd_m": 850,  # MarketsandMarkets 2025
    "ph_aqua_value_php_b": 124.02,  # BFAR 2023
    "ph_registered_aquafarmers": 260_961,  # BFAR FishR 2023
    "ph_sam_mid": 70_000,  # mobile-reachable finfish/shrimp (modeled)
    "fx_php_per_usd": 58.0,
}

PREMIUM_PRICE = 299
ENTERPRISE_PRICE = 1499

# --- Subscriber / funnel model (EOY and year-average for revenue) ---
YEARS = [1, 2, 3]

funnel = {
    1: {
        "free_eoy": 4_000,
        "premium_eoy": 150,
        "enterprise_eoy": 15,
        "premium_avg": 55,  # soft launch M8–12
        "enterprise_avg": 6,
        "marketplace": 25_000,
        "consultation": 10_000,
        "phase": "Phase 1 launch — free app + soft premium/enterprise pilots",
        "sam_penetration_paying_pct": None,  # filled later
    },
    2: {
        "free_eoy": 18_000,
        "premium_eoy": 1_400,
        "enterprise_eoy": 180,
        "premium_avg": 900,
        "enterprise_avg": 110,
        "marketplace": 450_000,
        "consultation": 120_000,
        "phase": "Phase 2 scale — BFAR/co-op channels + enterprise sales; breakeven year",
        "sam_penetration_paying_pct": None,
    },
    3: {
        "free_eoy": 32_000,
        "premium_eoy": 3_200,
        "enterprise_eoy": 350,
        "premium_avg": 2_500,
        "enterprise_avg": 280,
        "marketplace": 1_350_000,
        "consultation": 420_000,
        "phase": "Phase 2 maturity — retention, upsell, early Phase 3 prep",
        "sam_penetration_paying_pct": None,
    },
}

# --- Lean cost model (PHP) aimed at Y2 breakeven ---
costs = {
    1: {
        "setup": 450_000,  # legal, brand, app store, discovery
        "salaries": 1_440_000,  # 2 founders lean draw + 1 junior (avg ~P120k/mo team)
        "rent_ops": 180_000,
        "cloud_ai": 240_000,
        "marketing": 280_000,  # organic / community first
        "software_tools": 72_000,
        "legal_admin": 120_000,
        "variable_cogs": 60_000,  # payment fees + free-user serve
    },
    2: {
        "setup": 120_000,
        "salaries": 2_640_000,  # 4 FTEs
        "rent_ops": 280_000,
        "cloud_ai": 720_000,
        "marketing": 780_000,
        "software_tools": 96_000,
        "legal_admin": 160_000,
        "variable_cogs": 380_000,
    },
    3: {
        "setup": 80_000,
        "salaries": 4_200_000,  # 6–7 FTEs
        "rent_ops": 420_000,
        "cloud_ai": 1_320_000,
        "marketing": 1_200_000,
        "software_tools": 144_000,
        "legal_admin": 220_000,
        "variable_cogs": 780_000,
    },
}


def peso(n: float) -> str:
    if n < 0:
        return f"(P{abs(n):,.0f})"
    return f"P{n:,.0f}"


def build_pl():
    rows = []
    for y in YEARS:
        f = funnel[y]
        c = costs[y]
        prem_rev = f["premium_avg"] * PREMIUM_PRICE * 12
        ent_rev = f["enterprise_avg"] * ENTERPRISE_PRICE * 12
        total_rev = prem_rev + ent_rev + f["marketplace"] + f["consultation"]
        total_exp = sum(c.values())
        profit = total_rev - total_exp
        paying_eoy = f["premium_eoy"] + f["enterprise_eoy"]
        sam_pct = paying_eoy / MARKET["ph_sam_mid"] * 100
        f["sam_penetration_paying_pct"] = sam_pct
        rows.append(
            {
                "year": y,
                "phase": f["phase"],
                "free_eoy": f["free_eoy"],
                "premium_eoy": f["premium_eoy"],
                "enterprise_eoy": f["enterprise_eoy"],
                "premium_avg": f["premium_avg"],
                "enterprise_avg": f["enterprise_avg"],
                "prem_rev": prem_rev,
                "ent_rev": ent_rev,
                "marketplace": f["marketplace"],
                "consultation": f["consultation"],
                "total_rev": total_rev,
                "costs": c,
                "total_exp": total_exp,
                "profit": profit,
                "margin_pct": profit / total_rev * 100 if total_rev else 0,
                "sam_paying_pct": sam_pct,
                "som_free_pct": f["free_eoy"] / MARKET["ph_sam_mid"] * 100,
            }
        )
    return rows


def make_charts(pl):
    years = [f"Year {r['year']}" for r in pl]
    rev = [r["total_rev"] / 1e6 for r in pl]
    exp = [r["total_exp"] / 1e6 for r in pl]
    profit = [r["profit"] / 1e6 for r in pl]

    fig, axes = plt.subplots(2, 2, figsize=(14, 10), dpi=140)
    fig.suptitle(
        "AquaVision 3-Year Projection — Freemium (P299) + Enterprise (P1,499)\n"
        "Target: Year 2 Breakeven  |  Grounded in PH SAM ~70,000 aquafarmers",
        fontsize=13,
        fontweight="bold",
        color="#1B4965",
    )
    fig.patch.set_facecolor("#F8FAFC")

    # 1) Revenue vs Expenses vs Profit
    ax = axes[0, 0]
    x = range(len(years))
    w = 0.28
    ax.bar([i - w for i in x], rev, w, label="Revenue", color="#2A9D8F")
    ax.bar(x, exp, w, label="Expenses", color="#E76F51")
    ax.bar([i + w for i in x], profit, w, label="Profit/(Loss)", color="#1B4965")
    ax.axhline(0, color="#333", linewidth=0.8)
    ax.set_xticks(list(x))
    ax.set_xticklabels(years)
    ax.set_ylabel("PHP (millions)")
    ax.set_title("P&L — Revenue, Expenses, Profit")
    ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.3)

    # annotate BE
    ax.annotate(
        "Breakeven",
        xy=(1, profit[1]),
        xytext=(1.15, profit[1] + 1.2),
        fontsize=8,
        color="#2A9D8F",
        arrowprops=dict(arrowstyle="->", color="#2A9D8F"),
    )

    # 2) Revenue mix stacked
    ax = axes[0, 1]
    prem = [r["prem_rev"] / 1e6 for r in pl]
    ent = [r["ent_rev"] / 1e6 for r in pl]
    mkt = [r["marketplace"] / 1e6 for r in pl]
    con = [r["consultation"] / 1e6 for r in pl]
    ax.bar(years, prem, label="Premium P299", color="#457B9D")
    ax.bar(years, ent, bottom=prem, label="Enterprise P1,499", color="#1D3557")
    bottom2 = [a + b for a, b in zip(prem, ent)]
    ax.bar(years, mkt, bottom=bottom2, label="Marketplace", color="#A8DADC")
    bottom3 = [a + b for a, b in zip(bottom2, mkt)]
    ax.bar(years, con, bottom=bottom3, label="Consultations", color="#F4A261")
    ax.set_ylabel("PHP (millions)")
    ax.set_title("Revenue Mix by Stream")
    ax.legend(fontsize=7, loc="upper left")
    ax.grid(axis="y", alpha=0.3)

    # 3) Subscriber growth
    ax = axes[1, 0]
    free = [r["free_eoy"] / 1000 for r in pl]
    prem_n = [r["premium_eoy"] for r in pl]
    ent_n = [r["enterprise_eoy"] for r in pl]
    ax.plot(years, free, "o-", color="#2A9D8F", label="Free users EOY (thousands)")
    ax.set_ylabel("Free users (thousands)", color="#2A9D8F")
    ax2 = ax.twinx()
    ax2.plot(years, prem_n, "s-", color="#457B9D", label="Premium EOY")
    ax2.plot(years, ent_n, "^-", color="#1D3557", label="Enterprise EOY")
    ax2.set_ylabel("Paying accounts EOY")
    ax.set_title("User Funnel Growth (EOY)")
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, fontsize=7, loc="upper left")
    ax.grid(axis="y", alpha=0.3)

    # 4) SAM penetration + cumulative cash
    ax = axes[1, 1]
    sam = [r["sam_paying_pct"] for r in pl]
    cum = []
    running = 0
    for r in pl:
        running += r["profit"]
        cum.append(running / 1e6)
    ax.bar(years, sam, color="#A8DADC", label="Paying / PH SAM %")
    ax.set_ylabel("% of PH SAM (70k) as paying")
    ax2 = ax.twinx()
    ax2.plot(years, cum, "o-", color="#E76F51", linewidth=2, label="Cumulative P&L (P M)")
    ax2.axhline(0, color="#E76F51", linestyle="--", alpha=0.5)
    ax2.set_ylabel("Cumulative profit/(loss) P millions")
    ax.set_title("SOM vs SAM + Cumulative Cash Impact")
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, fontsize=7, loc="upper left")
    ax.grid(axis="y", alpha=0.3)

    plt.tight_layout(rect=[0, 0.02, 1, 0.93])
    path = OUT / "AquaVision-3Year-Freemium-Charts.png"
    fig.savefig(path, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    return path


def build_unit_economics(pl):
    """Derive per-account metrics from projection assumptions."""
    MONTHLY_CHURN = 0.05  # modeled midpoint (unit-economics.json)
    rows = []
    for i, r in enumerate(pl):
        prev = pl[i - 1] if i > 0 else None
        paying_avg = r["premium_avg"] + r["enterprise_avg"]
        paying_eoy = r["premium_eoy"] + r["enterprise_eoy"]
        c = r["costs"]
        blended_arpu_mo = (
            (r["premium_avg"] * PREMIUM_PRICE + r["enterprise_avg"] * ENTERPRISE_PRICE) / paying_avg
            if paying_avg
            else 0
        )
        sub_rev = r["prem_rev"] + r["ent_rev"]
        gross_margin = (sub_rev - c["variable_cogs"]) / sub_rev if sub_rev else 0
        prev_paying = (prev["premium_eoy"] + prev["enterprise_eoy"]) if prev else 0
        net_new_paying = max(paying_eoy - prev_paying, 1)
        cac = c["marketing"] / net_new_paying
        ltv = blended_arpu_mo * gross_margin / MONTHLY_CHURN
        cost_serve_free_mo = (c["cloud_ai"] + c["variable_cogs"] * 0.55) / 12 / max(r["free_eoy"], 1)
        contrib_mo = blended_arpu_mo * gross_margin
        payback_mo = cac / contrib_mo if contrib_mo else 0
        conversion = paying_eoy / max(r["free_eoy"], 1) * 100
        rows.append(
            {
                "year": r["year"],
                "prem_arpu_mo": PREMIUM_PRICE,
                "ent_arpu_mo": ENTERPRISE_PRICE,
                "blended_arpu_mo": blended_arpu_mo,
                "gross_margin": gross_margin,
                "cac": cac,
                "ltv": ltv,
                "ltv_cac": ltv / cac if cac else 0,
                "cost_serve_free_mo": cost_serve_free_mo,
                "payback_mo": payback_mo,
                "conversion_pct": conversion,
            }
        )
    return rows


def make_unit_economics_chart(pl):
    """Unit Economics chart for the Assumptions screen (course upload)."""
    ue = build_unit_economics(pl)
    years = [f"Year {u['year']}" for u in ue]

    fig, axes = plt.subplots(2, 2, figsize=(14, 10), dpi=140)
    fig.patch.set_facecolor("#F8FAFC")
    fig.suptitle(
        "AquaVision — Unit Economics (Assumptions Screen)\n"
        f"Premium P{PREMIUM_PRICE}/mo  |  Enterprise P{ENTERPRISE_PRICE}/mo  |  "
        "Modeled 5% monthly churn  |  Currency: PHP",
        fontsize=13,
        fontweight="bold",
        color="#1B4965",
    )

    # 1) ARPU by tier
    ax = axes[0, 0]
    x = range(len(years))
    w = 0.25
    prem = [u["prem_arpu_mo"] for u in ue]
    ent = [u["ent_arpu_mo"] for u in ue]
    blended = [u["blended_arpu_mo"] for u in ue]
    ax.bar([i - w for i in x], prem, w, label=f"Premium P{PREMIUM_PRICE}", color="#457B9D")
    ax.bar(x, ent, w, label=f"Enterprise P{ENTERPRISE_PRICE}", color="#1D3557")
    ax.bar([i + w for i in x], blended, w, label="Blended (weighted)", color="#2A9D8F")
    ax.set_xticks(list(x))
    ax.set_xticklabels(years)
    ax.set_ylabel("PHP / account / month")
    ax.set_title("Average Revenue Per Paying Account (ARPU)")
    ax.legend(fontsize=7)
    ax.grid(axis="y", alpha=0.3)

    # 2) LTV vs CAC
    ax = axes[0, 1]
    ltv = [u["ltv"] for u in ue]
    cac = [u["cac"] for u in ue]
    ax.bar([i - w / 2 for i in x], ltv, w, label="LTV (modeled)", color="#2A9D8F")
    ax.bar([i + w / 2 for i in x], cac, w, label="CAC (marketing / net-new paying)", color="#E76F51")
    for i, u in enumerate(ue):
        ax.annotate(
            f"{u['ltv_cac']:.1f}x",
            (i, max(ltv[i], cac[i]) + 200),
            ha="center",
            fontsize=8,
            color="#1B4965",
            fontweight="bold",
        )
    ax.set_xticks(list(x))
    ax.set_xticklabels(years)
    ax.set_ylabel("PHP per paying account")
    ax.set_title("Lifetime Value vs Customer Acquisition Cost")
    ax.legend(fontsize=7)
    ax.grid(axis="y", alpha=0.3)

    # 3) Cost to serve free user + gross margin
    ax = axes[1, 0]
    free_cost = [u["cost_serve_free_mo"] for u in ue]
    margin = [u["gross_margin"] * 100 for u in ue]
    ax.bar(years, free_cost, color="#A8DADC", label="Cost to serve / free user / mo")
    ax.set_ylabel("PHP / free user / month", color="#1B4965")
    ax2 = ax.twinx()
    ax2.plot(years, margin, "o-", color="#E76F51", linewidth=2, label="Subscription gross margin %")
    ax2.set_ylabel("Gross margin on subscriptions (%)", color="#E76F51")
    ax.set_title("Free-Tier Serve Cost & Paying Gross Margin")
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, fontsize=7, loc="upper right")
    ax.grid(axis="y", alpha=0.3)

    # 4) Summary metrics table
    ax = axes[1, 1]
    ax.axis("off")
    table_data = [
        ["Metric", "Year 1", "Year 2 (BE)", "Year 3"],
        ["Blended ARPU / mo", *[f"P{u['blended_arpu_mo']:,.0f}" for u in ue]],
        ["Gross margin (sub)", *[f"{u['gross_margin']*100:.0f}%" for u in ue]],
        ["CAC (per new paying)", *[f"P{u['cac']:,.0f}" for u in ue]],
        ["LTV (5% churn)", *[f"P{u['ltv']:,.0f}" for u in ue]],
        ["LTV : CAC", *[f"{u['ltv_cac']:.1f}x" for u in ue]],
        ["Payback (months)", *[f"{u['payback_mo']:.1f}" for u in ue]],
        ["Paying / free EOY", *[f"{u['conversion_pct']:.1f}%" for u in ue]],
    ]
    table = ax.table(
        cellText=table_data,
        loc="center",
        bbox=[0.02, 0.08, 0.96, 0.82],
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.5)
    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor("#CCCCCC")
        cell.set_linewidth(0.5)
        if r == 0:
            cell.set_facecolor("#1B4965")
            cell.set_text_props(color="white", fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor("#FAFAFA")
        if c == 0:
            cell.set_text_props(ha="left", fontweight="bold")
    ax.set_title("Unit Economics Summary", fontsize=11, color="#1B4965", pad=12)

    plt.tight_layout(rect=[0, 0.02, 1, 0.90])
    path = OUT / "AquaVision-3Year-Freemium-Unit-Economics.png"
    fig.savefig(path, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    return path


def make_revenue_net_income_chart(pl):
    """Revenue vs Net Income chart for the Assumptions screen (course upload)."""
    years = [f"Year {r['year']}" for r in pl]
    revenue = [r["total_rev"] for r in pl]
    net_income = [r["profit"] for r in pl]

    fig, ax = plt.subplots(figsize=(12, 7), dpi=140)
    fig.patch.set_facecolor("#F8FAFC")

    fig.suptitle(
        "AquaVision — Revenue vs. Net Income (Assumptions Screen)",
        fontsize=16,
        fontweight="bold",
        color="#1B4965",
        y=0.98,
    )
    fig.text(
        0.5,
        0.915,
        f"3-Year Freemium Projection  |  Premium P{PREMIUM_PRICE}/mo  |  Enterprise P{ENTERPRISE_PRICE}/mo  |  "
        "Target: Year 2 Breakeven  |  Currency: PHP",
        ha="center",
        fontsize=9,
        color="#555555",
    )

    x = range(len(years))
    width = 0.35
    bars_rev = ax.bar(
        [i - width / 2 for i in x],
        revenue,
        width,
        label="Total Revenue",
        color="#2A9D8F",
    )
    colors = ["#C0392B" if p < 0 else "#2E7D32" for p in net_income]
    bars_ni = ax.bar(
        [i + width / 2 for i in x],
        net_income,
        width,
        label="Net Income / (Loss)",
        color=colors,
    )

    for bar in bars_rev:
        h = bar.get_height()
        ax.annotate(
            f"P{h / 1e6:.2f}M",
            (bar.get_x() + bar.get_width() / 2, h),
            ha="center",
            va="bottom",
            fontsize=9,
            fontweight="bold",
            color="#1B4965",
        )
    for bar, p in zip(bars_ni, net_income):
        offset = max(abs(v) for v in net_income) * 0.03
        ax.annotate(
            f"P{p / 1e6:+.2f}M",
            (bar.get_x() + bar.get_width() / 2, p + (offset if p >= 0 else -offset)),
            ha="center",
            va="bottom" if p >= 0 else "top",
            fontsize=9,
            fontweight="bold",
            color="#C0392B" if p < 0 else "#2E7D32",
        )

    ax.axhline(0, color="#333333", linewidth=1)
    ax.set_xticks(list(x))
    ax.set_xticklabels(years, fontsize=11)
    ax.set_ylabel("Philippine Peso (PHP)")
    ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda v, _: f"P{v / 1e6:.0f}M"))
    ax.set_title("Revenue vs. Net Income — Path to Year 2 Breakeven", fontsize=12, color="#1B4965")
    ax.legend(loc="upper left", fontsize=9)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.set_facecolor("#FFFFFF")

    line_ax = ax.twinx()
    line_ax.plot(
        list(x),
        [r / 1e6 for r in revenue],
        color="#2A9D8F",
        marker="o",
        linewidth=2,
        alpha=0.35,
        label="_nolegend_",
    )
    line_ax.plot(
        list(x),
        [r / 1e6 for r in net_income],
        color="#1B4965",
        marker="s",
        linewidth=2,
        alpha=0.5,
        label="_nolegend_",
    )
    line_ax.set_yticks([])
    line_ax.set_ylim(ax.get_ylim())

    ax.annotate(
        "Breakeven",
        xy=(1, net_income[1]),
        xytext=(1.35, net_income[1] + max(revenue) * 0.12),
        fontsize=9,
        color="#2E7D32",
        fontweight="bold",
        arrowprops=dict(arrowstyle="->", color="#2E7D32"),
    )

    footer = "  |  ".join(
        f"Y{r['year']}: Rev {peso(r['total_rev'])} · Net {peso(r['profit'])}"
        for r in pl
    )
    fig.text(
        0.5,
        0.02,
        footer,
        ha="center",
        fontsize=8,
        color="#555555",
        bbox=dict(boxstyle="round,pad=0.35", facecolor="#E8F4FD", edgecolor="#1B4965"),
    )

    plt.tight_layout(rect=[0, 0.05, 1, 0.88])
    path = OUT / "AquaVision-3Year-Freemium-Revenue-vs-Net-Income.png"
    fig.savefig(path, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    return path


def make_summary_png(pl):
    """Single Assumptions-style summary image for course upload."""
    fig, ax = plt.subplots(figsize=(16, 11), dpi=130)
    ax.axis("off")
    fig.patch.set_facecolor("#F8FAFC")

    ax.text(
        0.5,
        0.97,
        "AquaVision — 3-Year Freemium Financial Projection (Assumptions View)",
        ha="center",
        va="top",
        fontsize=15,
        fontweight="bold",
        color="#1B4965",
        transform=ax.transAxes,
    )
    ax.text(
        0.5,
        0.935,
        "Free tier  |  Premium P299/mo  |  Enterprise P1,499/mo  |  Target: Year 2 Breakeven  |  Currency: PHP",
        ha="center",
        va="top",
        fontsize=9,
        color="#555",
        transform=ax.transAxes,
    )

    # Market box
    market_txt = (
        f"TAM: ~USD 1–2B aqua digital software (global)  ·  "
        f"Precision aqua tech USD 0.85B (MarketsandMarkets 2025)\n"
        f"SAM (PH): ~70,000 mobile-reachable finfish/shrimp aquafarmers "
        f"(derived from BFAR FishR 260,961 aquafarmers)\n"
        f"SOM Y3: {pl[2]['premium_eoy'] + pl[2]['enterprise_eoy']:,} paying "
        f"({pl[2]['sam_paying_pct']:.1f}% of SAM) + {pl[2]['free_eoy']:,} free "
        f"({pl[2]['som_free_pct']:.0f}% of SAM)"
    )
    ax.text(
        0.5,
        0.88,
        market_txt,
        ha="center",
        va="top",
        fontsize=8,
        color="#1B4965",
        transform=ax.transAxes,
        bbox=dict(boxstyle="round,pad=0.5", facecolor="#E8F4FD", edgecolor="#1B4965"),
    )

    headers = [
        "Metric",
        "Year 1",
        "Year 2 (BE)",
        "Year 3",
    ]
    table_data = [
        headers,
        ["Phase focus", "Launch free + soft paid", "Scale + Breakeven", "Profit & retention"],
        [
            "Free users (EOY)",
            f"{pl[0]['free_eoy']:,}",
            f"{pl[1]['free_eoy']:,}",
            f"{pl[2]['free_eoy']:,}",
        ],
        [
            "Premium avg / EOY",
            f"{pl[0]['premium_avg']} / {pl[0]['premium_eoy']}",
            f"{pl[1]['premium_avg']} / {pl[1]['premium_eoy']}",
            f"{pl[2]['premium_avg']} / {pl[2]['premium_eoy']}",
        ],
        [
            "Enterprise avg / EOY",
            f"{pl[0]['enterprise_avg']} / {pl[0]['enterprise_eoy']}",
            f"{pl[1]['enterprise_avg']} / {pl[1]['enterprise_eoy']}",
            f"{pl[2]['enterprise_avg']} / {pl[2]['enterprise_eoy']}",
        ],
        [
            "Premium revenue",
            peso(pl[0]["prem_rev"]),
            peso(pl[1]["prem_rev"]),
            peso(pl[2]["prem_rev"]),
        ],
        [
            "Enterprise revenue",
            peso(pl[0]["ent_rev"]),
            peso(pl[1]["ent_rev"]),
            peso(pl[2]["ent_rev"]),
        ],
        [
            "Marketplace + consult",
            peso(pl[0]["marketplace"] + pl[0]["consultation"]),
            peso(pl[1]["marketplace"] + pl[1]["consultation"]),
            peso(pl[2]["marketplace"] + pl[2]["consultation"]),
        ],
        [
            "TOTAL REVENUE",
            peso(pl[0]["total_rev"]),
            peso(pl[1]["total_rev"]),
            peso(pl[2]["total_rev"]),
        ],
        [
            "TOTAL EXPENSES",
            peso(pl[0]["total_exp"]),
            peso(pl[1]["total_exp"]),
            peso(pl[2]["total_exp"]),
        ],
        [
            "NET PROFIT / (LOSS)",
            peso(pl[0]["profit"]),
            peso(pl[1]["profit"]),
            peso(pl[2]["profit"]),
        ],
        [
            "Paying as % of PH SAM",
            f"{pl[0]['sam_paying_pct']:.2f}%",
            f"{pl[1]['sam_paying_pct']:.2f}%",
            f"{pl[2]['sam_paying_pct']:.2f}%",
        ],
    ]

    table = ax.table(
        cellText=table_data,
        loc="center",
        bbox=[0.05, 0.18, 0.9, 0.58],
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.5)

    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor("#CCCCCC")
        cell.set_linewidth(0.5)
        if r == 0:
            cell.set_facecolor("#1B4965")
            cell.set_text_props(color="white", fontweight="bold")
        elif table_data[r][0] in ("TOTAL REVENUE", "TOTAL EXPENSES", "NET PROFIT / (LOSS)"):
            cell.set_facecolor("#FFF3CD")
            cell.set_text_props(fontweight="bold")
            if table_data[r][0] == "NET PROFIT / (LOSS)":
                val = pl[c - 1]["profit"] if c > 0 else 0
                if c > 0:
                    cell.set_facecolor("#E8F5E9" if val >= 0 else "#FFEBEE")
        elif r % 2 == 0:
            cell.set_facecolor("#FAFAFA")
        if c == 0:
            cell.set_text_props(ha="left", fontweight="bold" if r > 0 else "bold")

    # Cost assumptions footer
    cost_lines = []
    for y, r in enumerate(pl, 1):
        c = r["costs"]
        cost_lines.append(
            f"Y{y} costs: salaries {peso(c['salaries'])} · cloud/AI {peso(c['cloud_ai'])} · "
            f"marketing {peso(c['marketing'])} · other {peso(r['total_exp'] - c['salaries'] - c['cloud_ai'] - c['marketing'])}"
        )
    ax.text(
        0.5,
        0.12,
        "\n".join(cost_lines),
        ha="center",
        va="top",
        fontsize=7.5,
        family="monospace",
        color="#333",
        transform=ax.transAxes,
    )
    ax.text(
        0.5,
        0.02,
        "Modeled projection · BFAR FishR 2023 · FAO SOFIA 2024 · MarketsandMarkets Precision Aquaculture 2025 · "
        "Y2 breakeven via lean opex + enterprise ARPU mix",
        ha="center",
        va="bottom",
        fontsize=7,
        color="#777",
        transform=ax.transAxes,
    )

    path = OUT / "AquaVision-3Year-Freemium-Assumptions.png"
    fig.savefig(path, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    return path


def make_pl_summary_png(pl):
    """Formatted PNG of the P&L Summary sheet (matches Excel tab)."""
    fig, ax = plt.subplots(figsize=(16, 12), dpi=130)
    ax.axis("off")
    fig.patch.set_facecolor("#F8FAFC")

    ax.text(
        0.5,
        0.97,
        "3-Year Profit & Loss (PHP)",
        ha="center",
        va="top",
        fontsize=16,
        fontweight="bold",
        color="#1B4965",
        transform=ax.transAxes,
    )
    ax.text(
        0.5,
        0.935,
        "AquaVision Freemium Model  |  Premium P299/mo  |  Enterprise P1,499/mo  |  Target: Year 2 Breakeven",
        ha="center",
        va="top",
        fontsize=9,
        color="#555",
        transform=ax.transAxes,
    )

    headers = ["Line Item", "Year 1", "Year 2", "Year 3"]
    pl_table = [headers]
    pl_lines = [
        ("Premium subscription revenue", "prem_rev", "money"),
        ("Enterprise subscription revenue", "ent_rev", "money"),
        ("Marketplace commission", "marketplace", "money"),
        ("Consultation fees", "consultation", "money"),
        ("Total Revenue", "total_rev", "total"),
        ("Total Expenses", "total_exp", "total"),
        ("Net Profit / (Loss)", "profit", "profit"),
        ("Profit Margin %", "margin_pct", "pct"),
        ("Paying accounts EOY", None, "count"),
        ("Paying as % of PH SAM", "sam_paying_pct", "sam_pct"),
    ]
    for label, key, fmt in pl_lines:
        row = [label]
        for r in pl:
            if fmt == "money":
                row.append(peso(r[key]))
            elif fmt == "total":
                row.append(peso(r[key]))
            elif fmt == "profit":
                row.append(peso(r[key]))
            elif fmt == "pct":
                row.append(f"{r[key]:.1f}%")
            elif fmt == "count":
                row.append(f"{r['premium_eoy'] + r['enterprise_eoy']:,}")
            elif fmt == "sam_pct":
                row.append(f"{r[key]:.2f}%")
        pl_table.append(row)

    table = ax.table(
        cellText=pl_table,
        loc="upper center",
        bbox=[0.05, 0.52, 0.9, 0.36],
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9)

    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor("#CCCCCC")
        cell.set_linewidth(0.5)
        if r == 0:
            cell.set_facecolor("#1B4965")
            cell.set_text_props(color="white", fontweight="bold")
        elif pl_table[r][0] in ("Total Revenue", "Total Expenses"):
            cell.set_facecolor("#FFF3CD")
            cell.set_text_props(fontweight="bold")
        elif pl_table[r][0] == "Net Profit / (Loss)":
            val = pl[c - 1]["profit"] if c > 0 else 0
            if c > 0:
                cell.set_facecolor("#E8F5E9" if val >= 0 else "#FFEBEE")
            cell.set_text_props(fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor("#FAFAFA")
        if c == 0:
            cell.set_text_props(ha="left", fontweight="bold" if r > 0 else "bold")

    ax.text(
        0.05,
        0.48,
        "EXPENSE DETAIL",
        ha="left",
        va="top",
        fontsize=11,
        fontweight="bold",
        color="#1B4965",
        transform=ax.transAxes,
    )

    cost_keys = list(costs[1].keys())
    exp_headers = ["Cost line"] + [f"Year {y}" for y in YEARS]
    exp_table = [exp_headers]
    for key in cost_keys:
        row = [key.replace("_", " ").title()]
        for y in YEARS:
            row.append(peso(costs[y][key]))
        exp_table.append(row)

    exp_tbl = ax.table(
        cellText=exp_table,
        loc="center",
        bbox=[0.05, 0.12, 0.9, 0.34],
        cellLoc="center",
    )
    exp_tbl.auto_set_font_size(False)
    exp_tbl.set_fontsize(8.5)

    for (r, c), cell in exp_tbl.get_celld().items():
        cell.set_edgecolor("#CCCCCC")
        cell.set_linewidth(0.5)
        if r == 0:
            cell.set_facecolor("#1B4965")
            cell.set_text_props(color="white", fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor("#FAFAFA")
        if c == 0:
            cell.set_text_props(ha="left", fontweight="bold" if r > 0 else "bold")

    ax.text(
        0.5,
        0.02,
        "Modeled projection · PH SAM ~70,000 aquafarmers · Y2 breakeven via lean opex + enterprise ARPU mix",
        ha="center",
        va="bottom",
        fontsize=7,
        color="#777",
        transform=ax.transAxes,
    )

    path = OUT / "AquaVision-3Year-Freemium-PL-Summary.png"
    fig.savefig(path, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    return path


def make_excel(pl):
    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1B4965")
    money_fill = PatternFill("solid", fgColor="FFF3CD")

    # --- Assumptions ---
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "AquaVision 3-Year Freemium Projection — Assumptions"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        f"Premium P{PREMIUM_PRICE}/mo · Enterprise P{ENTERPRISE_PRICE}/mo · "
        f"PH SAM {MARKET['ph_sam_mid']:,} · Target Year 2 Breakeven"
    )

    ws["A4"] = "PRICING"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = "Free tier"
    ws["B5"] = 0
    ws["A6"] = "Premium (farm account / month)"
    ws["B6"] = PREMIUM_PRICE
    ws["A7"] = "Enterprise (site / month)"
    ws["B7"] = ENTERPRISE_PRICE

    ws["A9"] = "MARKET ANCHORS"
    ws["A9"].font = Font(bold=True)
    anchors = [
        ("Global aqua digital TAM (USD B, midpoint)", MARKET["tam_software_usd_b"]),
        ("Precision aqua market 2025 (USD M, MnM)", MARKET["tam_precision_usd_m"]),
        ("PH aquaculture value (PHP B, BFAR 2023)", MARKET["ph_aqua_value_php_b"]),
        ("PH registered aquafarmers (BFAR FishR)", MARKET["ph_registered_aquafarmers"]),
        ("PH SAM — mobile finfish/shrimp (modeled)", MARKET["ph_sam_mid"]),
    ]
    for i, (label, val) in enumerate(anchors, 10):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    ws["A16"] = "FUNNEL DRIVERS"
    ws["A16"].font = Font(bold=True)
    headers = ["Driver", "Year 1", "Year 2", "Year 3"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=17, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    drivers = [
        ("Free users EOY", "free_eoy"),
        ("Premium avg (revenue)", "premium_avg"),
        ("Premium EOY", "premium_eoy"),
        ("Enterprise avg (revenue)", "enterprise_avg"),
        ("Enterprise EOY", "enterprise_eoy"),
        ("Marketplace revenue", "marketplace"),
        ("Consultation revenue", "consultation"),
    ]
    for i, (label, key) in enumerate(drivers, 18):
        ws.cell(row=i, column=1, value=label)
        for j, r in enumerate(pl):
            ws.cell(row=i, column=2 + j, value=r[key] if key in r else funnel[r["year"]][key])

    ws["A26"] = "COST ASSUMPTIONS (PHP)"
    ws["A26"].font = Font(bold=True)
    cost_keys = list(costs[1].keys())
    for col, h in enumerate(["Cost line"] + [f"Year {y}" for y in YEARS], 1):
        cell = ws.cell(row=27, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for i, key in enumerate(cost_keys, 28):
        ws.cell(row=i, column=1, value=key.replace("_", " ").title())
        for j, y in enumerate(YEARS):
            ws.cell(row=i, column=2 + j, value=costs[y][key]).number_format = '#,##0'

    ws.column_dimensions["A"].width = 42
    for col in "BCDE":
        ws.column_dimensions[col].width = 16

    # --- P&L ---
    ws2 = wb.create_sheet("P&L Summary")
    ws2["A1"] = "3-Year Profit & Loss (PHP)"
    ws2["A1"].font = Font(bold=True, size=14)
    pl_headers = ["Line Item", "Year 1", "Year 2", "Year 3"]
    for col, h in enumerate(pl_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    pl_lines = [
        ("Premium subscription revenue", "prem_rev"),
        ("Enterprise subscription revenue", "ent_rev"),
        ("Marketplace commission", "marketplace"),
        ("Consultation fees", "consultation"),
        ("Total Revenue", "total_rev"),
        ("Total Expenses", "total_exp"),
        ("Net Profit / (Loss)", "profit"),
        ("Profit Margin %", "margin_pct"),
        ("Paying accounts EOY", None),
        ("Paying as % of PH SAM", "sam_paying_pct"),
    ]
    for i, (label, key) in enumerate(pl_lines, 4):
        ws2.cell(row=i, column=1, value=label)
        for j, r in enumerate(pl):
            if key == "margin_pct":
                ws2.cell(row=i, column=2 + j, value=r[key] / 100).number_format = "0.0%"
            elif key is None:
                ws2.cell(
                    row=i,
                    column=2 + j,
                    value=r["premium_eoy"] + r["enterprise_eoy"],
                )
            elif key == "sam_paying_pct":
                ws2.cell(row=i, column=2 + j, value=r[key] / 100).number_format = "0.00%"
            else:
                cell = ws2.cell(row=i, column=2 + j, value=r[key])
                cell.number_format = '#,##0'
                if label in ("Total Revenue", "Total Expenses", "Net Profit / (Loss)"):
                    cell.fill = money_fill
                    cell.font = Font(bold=True)

    # Expense detail
    ws2.cell(row=16, column=1, value="EXPENSE DETAIL").font = Font(bold=True)
    for col, h in enumerate(["Cost line"] + [f"Year {y}" for y in YEARS], 1):
        cell = ws2.cell(row=17, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for i, key in enumerate(cost_keys, 18):
        ws2.cell(row=i, column=1, value=key.replace("_", " ").title())
        for j, y in enumerate(YEARS):
            ws2.cell(row=i, column=2 + j, value=costs[y][key]).number_format = '#,##0'

    ws2.column_dimensions["A"].width = 36
    for col in "BCD":
        ws2.column_dimensions[col].width = 16

    # Chart data sheet for openpyxl charts
    ws3 = wb.create_sheet("ChartData")
    ws3["A1"] = "Year"
    ws3["B1"] = "Revenue"
    ws3["C1"] = "Expenses"
    ws3["D1"] = "Profit"
    for i, r in enumerate(pl, 2):
        ws3.cell(row=i, column=1, value=f"Year {r['year']}")
        ws3.cell(row=i, column=2, value=r["total_rev"])
        ws3.cell(row=i, column=3, value=r["total_exp"])
        ws3.cell(row=i, column=4, value=r["profit"])

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Revenue vs Expenses vs Profit (PHP)"
    chart.y_axis.title = "PHP"
    chart.x_axis.title = "Year"
    data = Reference(ws3, min_col=2, min_row=1, max_col=4, max_row=4)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 10
    ws2.add_chart(chart, "F3")

    path = OUT / "AquaVision-3Year-Freemium-Projection.xlsx"
    wb.save(path)
    return path


def main():
    pl = build_pl()
    charts = make_charts(pl)
    unit_econ = make_unit_economics_chart(pl)
    rev_ni = make_revenue_net_income_chart(pl)
    summary = make_summary_png(pl)
    pl_summary = make_pl_summary_png(pl)
    xlsx = make_excel(pl)

    print("=== AquaVision 3-Year Freemium Projection ===")
    print(f"Premium: P{PREMIUM_PRICE}/mo | Enterprise: P{ENTERPRISE_PRICE}/mo")
    print(f"PH SAM: {MARKET['ph_sam_mid']:,}")
    print()
    for r in pl:
        print(
            f"Year {r['year']}: Rev {r['total_rev']:,.0f} | "
            f"Exp {r['total_exp']:,.0f} | Profit {r['profit']:,.0f} | "
            f"Paying EOY {r['premium_eoy']+r['enterprise_eoy']} "
            f"({r['sam_paying_pct']:.2f}% SAM)"
        )
    print()
    print(f"Charts:     {charts} ({charts.stat().st_size/1024:.0f} KB)")
    print(f"Unit Econ:  {unit_econ} ({unit_econ.stat().st_size/1024:.0f} KB)")
    print(f"Rev vs NI:  {rev_ni} ({rev_ni.stat().st_size/1024:.0f} KB)")
    print(f"Summary:    {summary} ({summary.stat().st_size/1024:.0f} KB)")
    print(f"P&L Summary:{pl_summary} ({pl_summary.stat().st_size/1024:.0f} KB)")
    print(f"Excel:      {xlsx}")
    return pl


if __name__ == "__main__":
    main()
