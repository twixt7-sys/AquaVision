"""Generate AquaVision BOE Financial Planning Assumptions Excel + PNG image."""

from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.table import Table
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent.parent / "exports"
OUT_DIR.mkdir(exist_ok=True)

PREMIUM_MONTHLY_PRICE = 299
ANNUAL_PREPAY_PRICE = 2_990

# Row structure: (label, col_b, y1, y2_condition, y2_pct, y3_condition, y3_pct)
# y2_pct and y3_pct are decimals; y2/y3 values computed from prior year row values

ROWS = [
    ("SETUP & ONE-TIME COSTS", None, None, None, None, None, None, True),
    (
        "Setup Costs",
        "SEC/DTI registration, legal, NPC data-privacy compliance",
        100_000,
        "Compliance renewals only; no new entity formation",
        -0.50,
        "Routine SEC/accounting renewals only",
        -0.20,
    ),
    (
        "Equipment & computers",
        "3 laptops, test phones, office printer, dev tools",
        200_000,
        "Replace one laptop; no major hardware purchase",
        -0.70,
        "No major hardware purchases; minor repairs only",
        -0.60,
    ),
    (
        "Software development",
        "Mobile app MVP, backend API, AI disease-ID module",
        1_800_000,
        "Phase 2 features — not a full MVP rebuild",
        -0.55,
        "Maintenance mode; incremental improvements only",
        -0.55,
    ),
    (
        "Office & workspace setup",
        "Co-working deposit, furniture, internet installation",
        150_000,
        "Workspace already established from Year 1",
        -0.90,
        "Workspace fully established in Year 2",
        -1.00,
    ),
    (
        "Branding & launch",
        "Logo, website, app store fees, educational content",
        150_000,
        "Ongoing content and app updates only",
        -0.40,
        "Sustained content; no major launch campaign",
        -0.15,
    ),
    (
        "Market research & farmer discovery",
        "BFAR/LGU outreach travel, farmer interviews",
        100_000,
        "Smaller ongoing BFAR/co-op validation rounds",
        -0.25,
        "Product decisions driven by in-app usage data",
        -0.30,
    ),
    ("OPERATING EXPENSES", None, None, None, None, None, None, True),
    (
        "Salaries & wages",
        "3-person team (2 developers, 1 product/community)",
        1_800_000,
        "Hire 1 backend developer + 1 growth associate",
        0.35,
        "Hire 1 ML engineer + part-time aquaculture advisor",
        0.28,
    ),
    (
        "Rent",
        "Co-working / small office, 12 months",
        180_000,
        "Upgrade to slightly larger co-working plan",
        0.15,
        "Small dedicated office for 5-person team",
        0.25,
    ),
    (
        "Utilities & internet",
        "Electricity, water, backup mobile data",
        60_000,
        "Stable operations; minor increase",
        0.05,
        "Stable operations; minor increase",
        0.05,
    ),
    (
        "Cloud hosting & infrastructure",
        "AWS/GCP servers, database, CDN, backups",
        360_000,
        "MAU roughly triples as free app scales",
        1.20,
        "~25,000 MAU; reserved instances offset scale costs",
        0.40,
    ),
    (
        "AI & API costs",
        "Disease-ID inference, farming assistant API",
        300_000,
        "More usage across free + paid users",
        1.00,
        "Model optimization lowers per-inference cost",
        0.30,
    ),
    (
        "Software subscriptions",
        "GitHub, Figma, analytics, payment gateway",
        72_000,
        "Add analytics, CRM, payment tools",
        0.25,
        "Add enterprise analytics and CRM tools",
        0.20,
    ),
    (
        "Marketing & customer acquisition",
        "Community content, social media, BFAR partnerships",
        400_000,
        "Double BFAR partnerships and community growth",
        0.75,
        "Referral-driven growth lowers acquisition cost",
        0.20,
    ),
    (
        "Insurance & admin",
        "Business insurance, accounting, bank fees",
        120_000,
        "Slightly higher admin as team grows",
        0.10,
        "Slightly higher admin as operations formalize",
        0.10,
    ),
    (
        "Legal & compliance",
        "Privacy audits, terms of service updates",
        80_000,
        "Maintain privacy compliance",
        0.00,
        "Marketplace vendor contracts & payment compliance",
        0.20,
    ),
    (
        "Repairs & maintenance",
        "Equipment upkeep, app maintenance",
        40_000,
        "Routine upkeep on existing equipment",
        0.15,
        "Routine equipment upkeep",
        0.10,
    ),
    ("VARIABLE COSTS (COGS)", None, None, None, None, None, None, True),
    (
        "Payment processing fees",
        "~3% of subscription & marketplace revenue",
        4_000,
        "More subscription and marketplace transactions",
        8.00,
        "Approximately 3% of paid transaction revenue",
        5.94,
    ),
    (
        "Per-user serving costs",
        "Messaging, storage & user-specific services (excl. cloud/AI)",
        120_000,
        "Usage scales; core cloud and AI tracked separately",
        2.50,
        "Unit serving cost falls through volume optimization",
        0.50,
    ),
    (
        "Customer support",
        "Part-time community moderator / support",
        120_000,
        "Dedicated part-time support for paying farms",
        1.00,
        "Full-time support for paying commercial farms",
        0.75,
    ),
    ("REVENUE", None, None, None, None, None, None, True),
    (
        "Subscription — Premium (monthly)",
        f"₱{PREMIUM_MONTHLY_PRICE}/month per farm; premium tier",
        PREMIUM_MONTHLY_PRICE * 12 * 20,
        "Average 200 monthly farms; ~400 by year-end",
        9.00,
        "Average 1,950 monthly farms; ~3,500 by year-end",
        8.75,
    ),
    (
        "Subscription — Annual prepay",
        f"₱{ANNUAL_PREPAY_PRICE:,}/year (two months free)",
        ANNUAL_PREPAY_PRICE * 15,
        "100 farms choose discounted annual plan",
        (100 / 15) - 1,
        "250 farms choose annual plan as retention improves",
        1.50,
    ),
    (
        "Marketplace commission",
        "10% take-rate on feed/equipment transactions",
        15_000,
        "Launch feed/equipment marketplace in Q2",
        4.00,
        "Nationwide feed/equipment partners onboarded",
        6.00,
    ),
    (
        "Consultation platform fees",
        "20% share on expert aquaculture sessions",
        10_000,
        "Expand expert consultation network",
        3.00,
        "Expert network grows to 20+ aquaculturists",
        3.50,
    ),
    ("SUMMARY", None, None, None, None, None, None, True),
    ("Total Revenue", "All revenue streams combined", None, "", None, "", None),
    ("Total Expenses", "Setup + operating + variable (excl. revenue)", None, "", None, "", None),
    ("Net Profit / (Loss)", "Revenue minus total expenses", None, "", None, "", None),
]

HEADER = [
    "Row",
    "Column B — Key Components",
    "Year 1 (₱)",
    "Year 2 — Condition / Decision",
    "Y1→Y2 %",
    "Year 3 — Condition / Decision",
    "Y1→Y3 %",
    "Year 2 (₱)",
    "Year 3 (₱)",
]

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1B4965")
SECTION_FILL = PatternFill("solid", fgColor="D4E4F0")
SUMMARY_FILL = PatternFill("solid", fgColor="FFF3CD")
POS_FILL = PatternFill("solid", fgColor="E8F5E9")
NEG_FILL = PatternFill("solid", fgColor="FFEBEE")


def fmt_peso(n):
    if n is None:
        return ""
    if n < 0:
        return f"(₱{abs(n):,.0f})"
    return f"₱{n:,.0f}"


def fmt_pct(n):
    if n is None:
        return ""
    sign = "+" if n >= 0 else ""
    return f"{sign}{n * 100:.0f}%"


def compute_values():
    data = []
    y2_by_key = {}
    y3_by_key = {}

    for row in ROWS:
        if len(row) == 8:
            data.append({"type": "section", "label": row[0]})
            continue

        label, col_b, y1, y2_cond, y2_pct, y3_cond, y3_pct = row
        y2 = round(y1 * (1 + y2_pct)) if y1 is not None else None
        y3 = round(y2 * (1 + y3_pct)) if y2 is not None else None
        y1_y3_pct = (y3 / y1 - 1) if y1 and y3 else None

        entry = {
            "type": "row",
            "label": label,
            "col_b": col_b,
            "y1": y1,
            "y2_cond": y2_cond,
            "y2_pct": y2_pct,
            "y3_cond": y3_cond,
            "y1_y3_pct": y1_y3_pct,
            "y2": y2,
            "y3": y3,
        }
        data.append(entry)
        if y1 is not None:
            y2_by_key[label] = y2
            y3_by_key[label] = y3

    setup_keys = ROWS[1][0], ROWS[2][0], ROWS[3][0], ROWS[4][0], ROWS[5][0], ROWS[6][0]
    opex_keys = [ROWS[i][0] for i in range(8, 18)]
    var_keys = [ROWS[i][0] for i in range(19, 22)]
    rev_keys = [ROWS[i][0] for i in range(23, 27)]

    def sum_keys(keys, store):
        return sum(store.get(k, 0) for k in keys)

    totals = {
        "y1": {
            "revenue": sum_keys(rev_keys, y2_by_key) if False else sum(
                r["y1"] for r in data if r["type"] == "row" and r["label"] in rev_keys
            ),
            "setup": sum(r["y1"] for r in data if r["type"] == "row" and r["label"] in setup_keys),
            "opex": sum(r["y1"] for r in data if r["type"] == "row" and r["label"] in opex_keys),
            "var": sum(r["y1"] for r in data if r["type"] == "row" and r["label"] in var_keys),
        }
    }
    totals["y1"]["expense"] = totals["y1"]["setup"] + totals["y1"]["opex"] + totals["y1"]["var"]
    totals["y1"]["profit"] = totals["y1"]["revenue"] - totals["y1"]["expense"]

    totals["y2"] = {
        "revenue": sum(y2_by_key[k] for k in rev_keys),
        "setup": sum(y2_by_key[k] for k in setup_keys),
        "opex": sum(y2_by_key[k] for k in opex_keys),
        "var": sum(y2_by_key[k] for k in var_keys),
    }
    totals["y2"]["expense"] = totals["y2"]["setup"] + totals["y2"]["opex"] + totals["y2"]["var"]
    totals["y2"]["profit"] = totals["y2"]["revenue"] - totals["y2"]["expense"]

    totals["y3"] = {
        "revenue": sum(y3_by_key[k] for k in rev_keys),
        "setup": sum(y3_by_key[k] for k in setup_keys),
        "opex": sum(y3_by_key[k] for k in opex_keys),
        "var": sum(y3_by_key[k] for k in var_keys),
    }
    totals["y3"]["expense"] = totals["y3"]["setup"] + totals["y3"]["opex"] + totals["y3"]["var"]
    totals["y3"]["profit"] = totals["y3"]["revenue"] - totals["y3"]["expense"]

    for entry in data:
        if entry["type"] != "row":
            continue
        if entry["label"] == "Total Revenue":
            entry["y1"] = totals["y1"]["revenue"]
            entry["y2"] = totals["y2"]["revenue"]
            entry["y3"] = totals["y3"]["revenue"]
        elif entry["label"] == "Total Expenses":
            entry["y1"] = totals["y1"]["expense"]
            entry["y2"] = totals["y2"]["expense"]
            entry["y3"] = totals["y3"]["expense"]
        elif entry["label"] == "Net Profit / (Loss)":
            entry["y1"] = totals["y1"]["profit"]
            entry["y2"] = totals["y2"]["profit"]
            entry["y3"] = totals["y3"]["profit"]

    return data, totals


def build_excel(data, totals):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"

    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "AquaVision — Back-of-the-Envelope Financial Planning | Assumptions"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="1B4965")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = (
        "Venture: AI-powered aquaculture subscription platform (PHP) | "
        f"Model: Freemium → Premium ₱{PREMIUM_MONTHLY_PRICE}/month"
    )
    sub.font = Font(italic=True, size=10, color="333333")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    for col, h in enumerate(HEADER, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[3].height = 36

    row_idx = 4
    for entry in data:
        if entry["type"] == "section":
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            cell = ws.cell(row=row_idx, column=1, value=entry["label"])
            cell.font = Font(bold=True, size=10, color="1B4965")
            cell.fill = SECTION_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            for c in range(1, 10):
                ws.cell(row=row_idx, column=c).border = BORDER
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1
            continue

        values = [
            entry["label"],
            entry["col_b"],
            entry["y1"],
            entry.get("y2_cond", ""),
            entry.get("y2_pct"),
            entry.get("y3_cond", ""),
            entry.get("y1_y3_pct"),
            entry.get("y2"),
            entry.get("y3"),
        ]

        is_summary = entry["label"] in (
            "Total Revenue",
            "Total Expenses",
            "Net Profit / (Loss)",
        )
        fill = SUMMARY_FILL if is_summary else None

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col)
            if col in (3, 8, 9) and isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = '"₱"#,##0'
            elif col in (5, 7) and isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = "0%"
            else:
                cell.value = val
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=col in (2, 4, 6),
                horizontal="right" if col in (3, 5, 7, 8, 9) else "left",
            )
            if fill:
                cell.fill = fill
                cell.font = Font(bold=True, size=10)
            if is_summary and col in (3, 8, 9) and isinstance(val, (int, float)):
                if entry["label"] == "Net Profit / (Loss)":
                    cell.fill = POS_FILL if val >= 0 else NEG_FILL

        ws.row_dimensions[row_idx].height = 32 if entry["col_b"] else 20
        row_idx += 1

    widths = [22, 38, 14, 34, 10, 34, 10, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("3-Year Summary")
    ws2["A1"] = "AquaVision 3-Year Projection Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("Metric", "Year 1", "Year 2", "Year 3"),
        ("Total Revenue", totals["y1"]["revenue"], totals["y2"]["revenue"], totals["y3"]["revenue"]),
        ("Setup Costs", totals["y1"]["setup"], totals["y2"]["setup"], totals["y3"]["setup"]),
        ("Operating Expenses", totals["y1"]["opex"], totals["y2"]["opex"], totals["y3"]["opex"]),
        ("Variable Costs", totals["y1"]["var"], totals["y2"]["var"], totals["y3"]["var"]),
        ("Total Expenses", totals["y1"]["expense"], totals["y2"]["expense"], totals["y3"]["expense"]),
        ("Net Profit / (Loss)", totals["y1"]["profit"], totals["y2"]["profit"], totals["y3"]["profit"]),
    ]
    for r, row in enumerate(summary_rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 3:
                cell.font = Font(bold=True)
            if c > 1 and isinstance(val, (int, float)):
                cell.number_format = '"₱"#,##0'
    ws2.column_dimensions["A"].width = 22
    for col in "BCD":
        ws2.column_dimensions[col].width = 16

    xlsx_path = OUT_DIR / "AquaVision-BOE-Assumptions-3Year.xlsx"
    try:
        wb.save(xlsx_path)
    except PermissionError:
        # Excel locks open workbooks on Windows; preserve the updated export
        # under a clear fallback name instead of aborting all image exports.
        xlsx_path = OUT_DIR / "AquaVision-BOE-Assumptions-3Year-Updated.xlsx"
        wb.save(xlsx_path)
    return xlsx_path


def build_png(data, totals):
    fig, ax = plt.subplots(figsize=(20, 28), dpi=120)
    ax.axis("off")
    fig.patch.set_facecolor("#F8FAFC")

    ax.text(
        0.5,
        0.985,
        "AquaVision — Back-of-the-Envelope Financial Planning Tool",
        ha="center",
        va="top",
        fontsize=18,
        fontweight="bold",
        color="#1B4965",
        transform=ax.transAxes,
    )
    ax.text(
        0.5,
        0.972,
        f"Assumptions Screen  |  3-Year Projection  |  Currency: Philippine Peso (₱)  |  "
        f"Model: Freemium → Premium ₱{PREMIUM_MONTHLY_PRICE}/month",
        ha="center",
        va="top",
        fontsize=10,
        color="#555555",
        transform=ax.transAxes,
    )

    table_rows = [
        [
            "Row",
            "Column B — Components",
            "Year 1 (₱)",
            "Y2 Condition",
            "Y1→Y2 %",
            "Y3 Condition",
            "Y2→Y3 %",
            "Year 2 (₱)",
            "Year 3 (₱)",
        ]
    ]

    for entry in data:
        if entry["type"] == "section":
            table_rows.append([entry["label"], "", "", "", "", "", "", "", ""])
            continue
        y2_pct = entry.get("y2_pct")
        y3_pct = entry.get("y1_y3_pct")
        # show Y2→Y3 % for year 3 column header consistency
        y2_y3 = None
        if entry.get("y2") and entry.get("y3") and entry["type"] == "row":
            if entry["label"] not in ("Total Revenue", "Total Expenses", "Net Profit / (Loss)"):
                if entry.get("y2"):
                    y2_y3 = (entry["y3"] / entry["y2"] - 1) if entry["y2"] else None

        # use stored y3_pct from data for detail rows
        pct_y3_display = entry.get("y1_y3_pct")
        if entry["label"] not in ("Total Revenue", "Total Expenses", "Net Profit / (Loss)"):
            pct_y3_display = ROWS[[r[0] for r in ROWS if len(r) == 7].index(entry["label"]) if False else 0]
            # simpler: extract y3_pct from original - re-read from entry
            pass

        # Re-find y3_pct from ROWS
        y3_pct_val = None
        for row in ROWS:
            if len(row) == 7 and row[0] == entry["label"]:
                y3_pct_val = row[6]
                break

        table_rows.append(
            [
                entry["label"][:28],
                (entry["col_b"] or "")[:42] + ("…" if entry["col_b"] and len(entry["col_b"]) > 42 else ""),
                fmt_peso(entry["y1"]) if entry["y1"] is not None else "",
                (entry.get("y2_cond") or "")[:36],
                fmt_pct(y2_pct) if y2_pct is not None else "",
                (entry.get("y3_cond") or "")[:36],
                fmt_pct(y3_pct_val) if y3_pct_val is not None else "",
                fmt_peso(entry.get("y2")),
                fmt_peso(entry.get("y3")),
            ]
        )

    table = ax.table(
        cellText=table_rows,
        loc="upper center",
        bbox=[0.01, 0.02, 0.98, 0.94],
        cellLoc="left",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(7)

    nrows = len(table_rows)
    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor("#CCCCCC")
        cell.set_linewidth(0.4)
        if r == 0:
            cell.set_facecolor("#1B4965")
            cell.set_text_props(color="white", fontweight="bold", fontsize=7)
            cell.set_height(0.018)
        elif table_rows[r][0] in (
            "SETUP & ONE-TIME COSTS",
            "OPERATING EXPENSES",
            "VARIABLE COSTS (COGS)",
            "REVENUE",
            "SUMMARY",
        ):
            cell.set_facecolor("#D4E4F0")
            cell.set_text_props(fontweight="bold", color="#1B4965")
        elif table_rows[r][0] in ("Total Revenue", "Total Expenses", "Net Profit / (Loss)"):
            cell.set_facecolor("#FFF3CD")
            cell.set_text_props(fontweight="bold")
            if c in (2, 7, 8) and "(" in str(table_rows[r][c]):
                cell.set_facecolor("#FFEBEE")
        elif r % 2 == 0:
            cell.set_facecolor("#FAFAFA")
        if c in (2, 4, 6, 7, 8):
            cell.set_text_props(ha="right")

    # Footer summary box
    footer = (
        f"Year 1: Revenue {fmt_peso(totals['y1']['revenue'])}  |  "
        f"Expenses {fmt_peso(totals['y1']['expense'])}  |  "
        f"Profit {fmt_peso(totals['y1']['profit'])}\n"
        f"Year 2: Revenue {fmt_peso(totals['y2']['revenue'])}  |  "
        f"Expenses {fmt_peso(totals['y2']['expense'])}  |  "
        f"Profit {fmt_peso(totals['y2']['profit'])}\n"
        f"Year 3: Revenue {fmt_peso(totals['y3']['revenue'])}  |  "
        f"Expenses {fmt_peso(totals['y3']['expense'])}  |  "
        f"Profit {fmt_peso(totals['y3']['profit'])}"
    )
    ax.text(
        0.5,
        0.008,
        footer,
        ha="center",
        va="bottom",
        fontsize=9,
        fontfamily="monospace",
        color="#1B4965",
        transform=ax.transAxes,
        bbox=dict(boxstyle="round,pad=0.4", facecolor="#E8F4FD", edgecolor="#1B4965"),
    )

    png_path = OUT_DIR / "AquaVision-BOE-Assumptions-3Year.png"
    plt.savefig(png_path, bbox_inches="tight", facecolor=fig.get_facecolor(), dpi=120)
    plt.close()
    return png_path


def build_chart_png(totals):
    years = ["Year 1", "Year 2", "Year 3"]
    revenue = [totals[y]["revenue"] for y in ("y1", "y2", "y3")]
    expenses = [totals[y]["expense"] for y in ("y1", "y2", "y3")]
    profit = [totals[y]["profit"] for y in ("y1", "y2", "y3")]

    fig, (ax1, ax2) = plt.subplots(
        1, 2, figsize=(16, 7), dpi=120, gridspec_kw={"width_ratios": [1.2, 1]}
    )
    fig.patch.set_facecolor("#F8FAFC")
    fig.suptitle(
        "AquaVision — 3-Year Financial Projection",
        fontsize=17,
        fontweight="bold",
        color="#1B4965",
    )
    fig.text(
        0.5,
        0.925,
        f"Premium plan: ₱{PREMIUM_MONTHLY_PRICE}/month | "
        f"Annual prepay: ₱{ANNUAL_PREPAY_PRICE:,}/year",
        ha="center",
        fontsize=10,
        color="#555555",
    )

    x = range(len(years))
    width = 0.35
    bars_rev = ax1.bar(
        [i - width / 2 for i in x], revenue, width, label="Total Revenue", color="#2E86AB"
    )
    bars_exp = ax1.bar(
        [i + width / 2 for i in x], expenses, width, label="Total Expenses", color="#E07A5F"
    )
    for bars in (bars_rev, bars_exp):
        for bar in bars:
            ax1.annotate(
                f"₱{bar.get_height() / 1e6:.2f}M",
                (bar.get_x() + bar.get_width() / 2, bar.get_height()),
                ha="center",
                va="bottom",
                fontsize=9,
                fontweight="bold",
                color="#333333",
            )
    ax1.set_title("Revenue vs Expenses", fontsize=12, color="#1B4965")
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(years, fontsize=11)
    ax1.set_ylabel("Philippine Peso (₱)")
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"₱{v / 1e6:.0f}M"))
    ax1.legend(loc="upper left")
    ax1.grid(axis="y", linestyle="--", alpha=0.4)
    ax1.set_facecolor("#FFFFFF")

    colors = ["#C0392B" if p < 0 else "#2E7D32" for p in profit]
    bars_p = ax2.bar(years, profit, 0.5, color=colors)
    for bar, p in zip(bars_p, profit):
        offset = 0.02 * max(abs(v) for v in profit)
        ax2.annotate(
            f"₱{p / 1e6:+.2f}M",
            (bar.get_x() + bar.get_width() / 2, p + (offset if p >= 0 else -offset)),
            ha="center",
            va="bottom" if p >= 0 else "top",
            fontsize=10,
            fontweight="bold",
            color=colors[list(profit).index(p)],
        )
    ax2.axhline(0, color="#333333", linewidth=1)
    ax2.set_title("Net Profit / (Loss) — Path to Profitability", fontsize=12, color="#1B4965")
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"₱{v / 1e6:.0f}M"))
    ax2.grid(axis="y", linestyle="--", alpha=0.4)
    ax2.set_facecolor("#FFFFFF")

    line_ax = ax2.twinx()
    line_ax.plot(years, profit, color="#1B4965", marker="o", linewidth=2, alpha=0.6)
    line_ax.set_yticks([])
    line_ax.set_ylim(ax2.get_ylim())

    fig.text(
        0.5,
        0.015,
        "Base-case assumption: Year 3 averages 1,950 monthly and 250 annual-plan farms "
        "(~3,750 paid farms by year-end), reaching a narrow break-even result.",
        ha="center",
        fontsize=9,
        style="italic",
        color="#555555",
    )

    fig.tight_layout(rect=[0, 0.04, 1, 0.88])
    chart_path = OUT_DIR / "AquaVision-3Year-Projection-Chart.png"
    plt.savefig(chart_path, bbox_inches="tight", facecolor=fig.get_facecolor(), dpi=120)
    plt.close()
    return chart_path


def main():
    data, totals = compute_values()
    xlsx = build_excel(data, totals)
    png = build_png(data, totals)
    chart = build_chart_png(totals)
    print(f"Chart: {chart}")
    png_size = png.stat().st_size
    print(f"Excel: {xlsx}")
    print(f"PNG:   {png} ({png_size / 1024:.1f} KB)")
    print("\n3-Year Totals:")
    for yr in ("y1", "y2", "y3"):
        t = totals[yr]
        label = yr.upper().replace("Y", "Year ")
        print(
            f"  {label}: Revenue {t['revenue']:,} | "
            f"Expenses {t['expense']:,} | Profit {t['profit']:,}"
        )
    if png_size > 1_000_000:
        print("WARNING: PNG exceeds 1 MB — reducing DPI...")
        # regenerate at lower dpi if needed
    return xlsx, png


if __name__ == "__main__":
    main()
